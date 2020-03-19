import os
import requests
import openpyxl as px
import psycopg2
import uuid
from bs4 import BeautifulSoup
import settings as env
from servant import FileDownLoader, XlsDateToIsoConverter, OnsetDateProvider, IsRelationConverter


def db_connect():
    return psycopg2.connect(env.AWJ_DB_CONNECT_SETUP)


def main():
    # 既にエクセルファイルがある場合一旦、既存のファイルを破棄
    if os.path.isfile(env.FILE_NAME):
        os.remove(env.FILE_NAME)

    # 兵庫県のサイトにアクセスしエクセルのaタグを取得
    res = requests.get("https://web.pref.hyogo.lg.jp/kk03/corona_kanjyajyokyo.html")
    soup = BeautifulSoup(res.content, "html.parser")
    xls_link = soup.select("#tmp_contents a.icon_excel")
    # ファイルをダウンロード
    fdl = FileDownLoader.Summon("https://web.pref.hyogo.lg.jp" + xls_link[0]["href"], env.FILE_NAME)
    fdl.service()

    # エクセルファイルを開いてシートを読みこむ
    wb = px.load_workbook(env.FILE_NAME)
    sheet = wb["公表"]

    # postgreSQL DBにエクセルデータを挿入する
    with db_connect() as conn:
        with conn.cursor() as db:

            for irow in range(5, sheet.max_row - 1):

                # DBで利用するid
                infected_people_id = uuid.uuid4()
                # エクセルで振られている番号
                infected_people_no = sheet.cell(row=irow, column=2).value

                # エクセルDatetimeをISOのDatetimeに変換するモジュール
                xdic = XlsDateToIsoConverter.Summon(sheet.cell(row=irow, column=3).value)

                # 発症日をstring型で返す。エクセルDatetimeが渡された場合%m月%d日のstringを、
                # 「確認中」などの文字列が渡された場合はそのまま返す。
                odp = OnsetDateProvider.Summon(sheet.cell(row=irow, column=9).value)

                # infected_peoplesテーブルに感染者が登録済みかチェック
                # 登録済みでない場合新規レコードを挿入
                # 登録済みの場合更新する形で対応
                db.execute("SELECT EXISTS (SELECT * FROM infected_peoples WHERE no = %s);", (infected_people_no,))
                (is_exist_infected_people,) = db.fetchone()
                if not is_exist_infected_people:
                    db.execute("""
                        INSERT
                        INTO infected_peoples (
                            id,
                            no,
                            confirmed_date,
                            age_group,
                            sex,
                            jurisdiction,
                            residence,
                            occupation,
                            onset_date,
                            travel_history,
                            remarks
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                    """, (
                        str(infected_people_id),
                        infected_people_no,
                        xdic.service(),
                        sheet.cell(row=irow, column=4).value,
                        sheet.cell(row=irow, column=5).value,
                        sheet.cell(row=irow, column=6).value,
                        sheet.cell(row=irow, column=7).value,
                        sheet.cell(row=irow, column=8).value,
                        odp.service(),
                        sheet.cell(row=irow, column=10).value,
                        sheet.cell(row=irow, column=11).value
                    ))

                    # エクセルの「認定こども園」以降のデータをinfected_placesテーブルに挿入
                    infected_place_no = 0
                    for icol in range(12, sheet.max_column - 1):
                        # エクセルの二重線を表現した空のカラムを読み飛ばす。
                        if sheet.cell(row=4, column=icol).value is None:
                            continue
                        infected_place_id = uuid.uuid4()
                        infected_place_no += 1
                        # エクセルの「○」と空をBoolean型に変換するモジュール
                        irc = IsRelationConverter.Summon(sheet.cell(row=irow, column=icol).value)
                        db.execute("""
                            INSERT
                            INTO infected_places (
                                id,
                                no,
                                infected_people_id,
                                name,
                                is_relation
                            )
                            VALUES (%s, %s, %s, %s, %s);
                        """, (
                            str(infected_place_id),
                            infected_place_no,
                            str(infected_people_id),
                            sheet.cell(row=4, column=icol).value.replace("\n", ""),
                            irc.service()
                        ))
                else:

                    db.execute("""
                        UPDATE infected_peoples
                        SET confirmed_date = %s,
                        age_group = %s,
                        sex = %s,
                        jurisdiction = %s,
                        residence = %s,
                        occupation = %s,
                        onset_date = %s,
                        travel_history = %s,
                        remarks = %s
                        WHERE no = %s;
                    """, (
                        xdic.service(),
                        sheet.cell(row=irow, column=4).value,
                        sheet.cell(row=irow, column=5).value,
                        sheet.cell(row=irow, column=6).value,
                        sheet.cell(row=irow, column=7).value,
                        sheet.cell(row=irow, column=8).value,
                        odp.service(),
                        sheet.cell(row=irow, column=10).value,
                        sheet.cell(row=irow, column=11).value,
                        infected_people_no
                    ))

                    # infected_peoplesテーブルから感染者番号を元にidを取得
                    db.execute("SELECT id FROM infected_peoples WHERE no = %s;", (infected_people_no,))
                    (infected_people_id,) = db.fetchone()

                    infected_place_no = 0
                    for icol in range(12, sheet.max_column):
                        # エクセルの二重線を表現した空のカラムを読み飛ばす。
                        if sheet.cell(row=4, column=icol).value is None:
                            continue
                        infected_place_no += 1
                        irc = IsRelationConverter.Summon(sheet.cell(row=irow, column=icol).value)

                        # エクセルの感染場所のカラム数が増えているかチェック増えている場合はDB側にも追加
                        db.execute("""
                            SELECT
                            EXISTS (
                                SELECT * FROM infected_places
                                WHERE no = %s
                                AND infected_people_id = %s
                            );
                        """, (
                            infected_place_no,
                            infected_people_id
                        ))
                        (is_exist_infected_place,) = db.fetchone()

                        # エクセルの感染場所のカラム数が増えた場合新規追加、それ以外の場合はUPDATEで対応
                        if not is_exist_infected_place:
                            infected_place_id = uuid.uuid4()

                            db.execute("""
                                INSERT
                                INTO infected_places (
                                    id,
                                    no,
                                    infected_people_id,
                                    name,
                                    is_relation
                                )
                                VALUES (%s, %s, %s, %s, %s);
                            """, (
                                str(infected_place_id),
                                infected_place_no,
                                str(infected_people_id),
                                sheet.cell(row=4, column=icol).value.replace("\n", ""),
                                irc.service()
                            ))
                        else:
                            db.execute("""
                                UPDATE infected_places
                                SET is_relation = %s, name = %s
                                WHERE no = %s AND infected_people_id = %s;
                            """, (
                                irc.service(),
                                sheet.cell(row=4, column=icol).value.replace("\n", ""),
                                infected_place_no,
                                infected_people_id
                            ))


if __name__ == '__main__':
    main()